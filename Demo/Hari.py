import pyodbc

class Hari:
    excelUrl = 'C:/Development/Robot/DotCom/Resources/TestData/qat01.xlsx'
    sheetName = 'REG'
    testcaseNo = 'TC12'
    searchHeaderName = 'Order_Id'

    # def update_excel_data_by_cell_name(self, excelUrl, sheetName, testcaseNo, searchHeaderName):
        #Parameters: (xlsx url location, Sheet Name, Test Case Number, required column value)
    # self.excelUrl = excelUrl
    # self.sheetName = sheetName
    # self.testcaseNo = testcaseNo
    # self.searchHeaderName = searchHeaderName
    import openpyxl

    # Open xlsx file for reading data
    wb = openpyxl.load_workbook(excelUrl)
    sheet = wb.get_sheet_by_name(sheetName)

    # Find Total count of max Rows and Columns
    rows = sheet.max_row
    columns = sheet.max_column

    # First Row data is a header data in the excel sheet.
    # Search header column location where we need to find out the value
    for y in range(1, columns + 1):
        header = sheet.cell(row=1, column=y).value
        if header == searchHeaderName:
            columnum = y
            break

    # Loop all rows from 2nd row (1st row is a header row) with idendified column to get right data.
    # Return search cata
    for x in range(2, rows + 1):
        tc_name = sheet.cell(row=x, column=1).value
        if tc_name == testcaseNo:
            sheet.cell(row=x, column=columnum).value = 12345
            wb.save(excelUrl)
            # return testdata














    # # clname = 'Region'
    # # reg = 'US'
    # # query = 'SELECT Region FROM DPI_Residential_Deactivate_ServiceOrders'
    # ordernum = 2222226
    # region = 'US'
    # conn_str = (
    #     r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
    #     r'DBQ=C:\Automation\TestData.mdb;'
    # )
    # cnxn = pyodbc.connect(conn_str)
    # crsr = cnxn.cursor()
    # crsr.execute("UPDATE DPI_Residential_Deactivate_ServiceOrders SET Order_Num = ? WHERE Region = ?", ordernum, region)
    # crsr.commit()
    # crsr.close()
    #
    #
    # # crsr.execute(query)
    # # print("Executed sql query")
    # # for row in crsr:
    # #     coldata = row.__getattribute__(clname)
    # #     if coldata == reg:
    # #         crsr.execute(" UPDATE DPI_Residential_Deactivate_ServiceOrders SET Order_Num = '222222' WHERE Region = 'US' ")
    # #         print("Executed update query")
    # #         crsr.commit()
    # # crsr.close()
    #
    #
    # # "UPDATE progress SET CockpitDrill = ? WHERE progress_primarykey = ?", newcockpitdrillvalue, oldprimarykeyvalue)
class readExcel:

    def get_address_data(self, excelUrl, sheetName, testcaseNo, searchHeaderName):
        #Parameters: (xlsx url location, Sheet Name, Test Case Number, required column value)
        self.excelUrl = excelUrl
        self.sheetName = sheetName
        self.testcaseNo = testcaseNo
        self.searchHeaderName = searchHeaderName
        import openpyxl

        #Open xlsx file for reading data
        wb = openpyxl.load_workbook(excelUrl)
        sheet = wb.get_sheet_by_name(sheetName)

        #Find Total count of max Rows and Columns
        rows = sheet.max_row
        columns = sheet.max_column

        #First Row data is a header data in the excel sheet.
        #Search header column location where we need to find out the value
        for y in range(1,columns+1):
            header = sheet.cell(row=1, column=y).value
            if header==searchHeaderName:
                columnum = y
                break

        # Loop all rows from 2nd row (1st row is a header row) with idendified column to get right data.
        #Return search cata
        for x in range(2, rows+1):
            tc_name = sheet.cell(row=x, column=1).value
            if tc_name == testcaseNo:
                testdata = sheet.cell(row=x, column=columnum).value
                #print (testdata)
                return testdata